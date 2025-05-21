import pandas as pd
import json

# df = pd.read_csv('company_data.csv', sep=",")

# # Write a code for the show starting 10 records
# starting_10_records = df.head(10) 
# print("Starting 10 Records")
# print(starting_10_records)


# # Write a code for the show last 5 records
# last_5_records = df.tail(5)
# print("Ending 5 Records")
# print(last_5_records)


# # Write a code for the show only company name and location
# # result = df.drop('CompanyID',axis='columns') #Delete single column
# columns_to_drop = ['CompanyID','Unit','Rate'] #Delete multiple columns
# result = df.drop(columns_to_drop, axis='columns')
# print(result)

# print(df)


# # Write a code the the show average unit
# average_unit = df['Unit'].mean()
# print('Average Unit', average_unit)

# # Write a code the the show average Rate
# average_rate = df['Rate'].mean()
# print('Average Rate', "{:.2f}".format(average_rate)) #two disit after dicimal


# # Total Unit and Rate of the data frame
# total_unit = df['Unit'].sum()
# total_rate = df['Rate'].sum()

# print('Total Unit', total_unit)
# print('Total Rate', '{:.2f}'.format(total_rate))


# # Write a code for the show to 10 rate
# df_sorted = df.sort_values(by='Rate',ascending=False)
# top_10_sorted = df_sorted.head(10)
# print(top_10_sorted)


# # Write a code for the show company name wise data
# df_sorted = df.sort_values(by='CompanyName', ascending=True)
# top_10_sorted_company = df_sorted.head(10)
# print(top_10_sorted_company)

# Write a code for the show companpany name start with k
# filtered_df = df[df['CompanyName'].str.lower().str.startswith('k')]
# print(filtered_df)


# filter_col = df['CompanyName'] #only show company data

# for i in df['CompanyName']:
#     if i.startswith('K'):
#         print(i)

# filtered_df = df[df['CompanyName'].str.contains('Leon', case=False)] # case=False for case-insensitive search
# print(filtered_df)


# #show 10 data without index
# print(df.head(10).to_string(index=False))

#Skip first row
# df = pd.read_csv('company_data.csv', skiprows=1)
# print(df)

# # Show all data of the those which unit more then 300
# result = df[df['Unit']>300]
# result = result.sort_values(by='Unit', ascending=True)

# print(result)

# # Show all data of the those which unit between 301 to 400
# result = df[df['Unit'].between(300,400)]
# result = result.sort_values(by='Unit', ascending=True)

# print(result)


# Show all data of the those which unit between 301 to 400 and sum of number
# result = df[df['Unit'].between(300,400)]
# unit_sum = result['Unit'].sum()
# print(unit_sum)

# unit_sum = df.loc[df['Unit'].between(300, 400), 'Unit'].sum()
# print(unit_sum)


# # Replace blank value to the 0
# result = df.head(10)
# result.fillna(0, inplace=True)
# print(result)


#---------------------------------------------------------------------------------------------------

# data = pd.read_json('company_row_data.json')

# data = open('company_row_data.json','r')
# json_string = data.read()
# # print(type(json_string))
# python_dict = json.loads(json_string) #convert json to dict

# # print(type(python_dict))

# for i in range(0,10):
#     print(python_dict[i])
#---------------------------------------------------------------------------------------------------

# df = pd.read_csv('employees.csv', sep=",")
# result = df.groupby('Department').sum()

# # Sum of the Salary Department wise
# result = df.groupby('Department')['Salary'].sum()
# print(result)


# # Average Salary of employee Department wise
# result = df.groupby('Department')['Salary'].mean()
# print(result)

#---------------------------------------------------------------------------------------------------
df = pd.read_csv('company_data.csv', sep=",")

# # CSV to Excel Conversion
# columns_list = ['Location','CompanyName','Rate']
# final_data = df[columns_list]

# final_data.to_excel('company_details.xlsx', index=False)

#---------------------------------------------------------------------------------------------------

# import openpyxl
# from openpyxl.styles import PatternFill, Font

# # Load the workbook and select the sheet
# wb = openpyxl.load_workbook("company_details.xlsx")
# ws = wb['Sheet1']

# # Define yellow fill (standard hex code: FFFF00)
# yellow_fill = PatternFill(start_color="FFFF00",
#                           fill_type="solid")

# red_font = Font(color="FF0000")

# # Apply the fill to cell A1
# cell = ws["A3"]
# cell.fill = yellow_fill
# cell.font = red_font

# # Save the modified workbook
# wb.save("output_data.xlsx")


#---------------------------------------------------------------------------------------------------
# Highlight Cell Background color
import openpyxl
from openpyxl.styles import PatternFill

# Open the Excel file
wb = openpyxl.load_workbook("company_details.xlsx")
ws = wb['Sheet1']

# Create yellow color fill
yellow_fill = PatternFill(fill_type='solid', start_color='FFFF00')

# Go through each row starting from row 2 (to skip the header)
for row in ws.iter_rows(min_row=2): #iterate each row skip first row start iterating from second row
    rate_cell = row[2]  # 3rd column is "Rate"
    if rate_cell.value > 400:
        rate_cell.fill = yellow_fill  # Apply yellow color

# # Set the height of row 2
# ws.row_dimensions[2].height = 30  # You can change 30 to whatever height you want       

ws.column_dimensions['B'].width = 30  # You can change 30 to any width you want(width 30 unit[1 unit ≈ 7-8 pixels])
ws.column_dimensions['A'].width = 30
# Save the file with a new name
wb.save("company_details_highlighted.xlsx")



